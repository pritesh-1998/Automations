import openpyxl
# Creation of workbook object
# openpyxl.workbook()
wb=openpyxl.load_workbook(r"C:\Users\plath\Documents\programming codes\Python automations\Automations\excel_operations\sheet.xlsx")

#print sheets name
print(wb.sheetnames)
data_sheet=wb["Sheet1"]

#accesing the cells
#here a=column and 1=row number
# cell=data_sheet["a1"]
# #print cell value
# print(cell.value)
# #changings thevalye of that cell
# cell.value="ID"
# print(cell.value)
# #print the value in that row
# print(cell.row)
# #print the value in that column
# print(cell.column)
# #print the exact coordinates of that cell
# print(cell.coordinate)

#getting al the values at a time
data_sheet=wb["Sheet1"]
cell=data_sheet["a1"]
# for row in range(1,data_sheet.max_row+1):
#     for column in range(1,data_sheet.max_column+1):
#         cell=data_sheet.cell(row,column)
#         print(cell.value)

#getting the value of specific row or column
column_data=data_sheet["a"]
print(column_data)

#getting al the values from row a to c
columns=data_sheet["a:c"]
print(columns)
#create sheet
# wb.create_sheet("sample",0)

